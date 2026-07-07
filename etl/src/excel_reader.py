"""Lectura de hojas Excel con encabezados multinivel (3 filas fusionadas).

BD_ENF_CAPIIIM_1.xlsx no tiene encabezados de una sola fila: cada columna
real se arma combinando 3 filas superiores (con celdas combinadas verticales
y horizontales, p.ej. en la hoja ANEMIA: fila1="1er Control", fila2="Enfermería",
fila3="Fecha"). openpyxl no “desfusiona” eso solo, así que este módulo:

1. Construye un mapa (fila, columna) -> valor para toda celda combinada,
   propagando el valor de la celda superior-izquierda a todo el rango.
2. Junta las `header_rows` filas de cada columna en un único nombre de
   columna "aplanado" (p.ej. "1er Control Enfermeria Fecha").
3. Devuelve un DataFrame con esos nombres de columna y una columna auxiliar
   `_excel_row` con el número de fila real del Excel (para trazabilidad en
   el reporte de descartados).
"""

from __future__ import annotations

import re
import unicodedata

import openpyxl
import pandas as pd


def normalizar_texto_busqueda(valor: object) -> str:
    """minúsculas, sin tildes, sin espacios/puntuación repetidos. Usado tanto
    para nombres de hoja como para nombres de columna aplanados, así los
    extractors pueden buscar candidatos sin preocuparse de mayúsculas/tildes."""
    if valor is None:
        return ""
    texto = str(valor).strip().lower()
    texto = unicodedata.normalize("NFKD", texto)
    texto = "".join(c for c in texto if not unicodedata.combining(c))
    texto = re.sub(r"[^a-z0-9]+", " ", texto)
    return re.sub(r"\s+", " ", texto).strip()


def encontrar_hoja(nombres_disponibles: list[str], candidatos: list[str]) -> str | None:
    """Busca, entre los nombres de hoja reales del libro, el que mejor matchea
    alguno de los `candidatos` (alias conocidos para esa hoja lógica)."""
    normalizados = {nombre: normalizar_texto_busqueda(nombre) for nombre in nombres_disponibles}
    for candidato in candidatos:
        candidato_norm = normalizar_texto_busqueda(candidato)
        for nombre, norm in normalizados.items():
            if norm == candidato_norm:
                return nombre
    for candidato in candidatos:
        candidato_norm = normalizar_texto_busqueda(candidato)
        for nombre, norm in normalizados.items():
            if candidato_norm in norm or norm in candidato_norm:
                return nombre
    return None


def encontrar_columna(columnas: list[str], candidatos: list[str]) -> str | None:
    """Encuentra, entre las columnas aplanadas de un DataFrame, la que mejor
    matchea alguno de los alias en `candidatos` (ver src/aliases.py)."""
    normalizadas = {col: normalizar_texto_busqueda(col) for col in columnas}
    for candidato in candidatos:
        candidato_norm = normalizar_texto_busqueda(candidato)
        for col, norm in normalizadas.items():
            if norm == candidato_norm:
                return col
    for candidato in candidatos:
        candidato_norm = normalizar_texto_busqueda(candidato)
        if not candidato_norm:
            continue
        for col, norm in normalizadas.items():
            if candidato_norm in norm:
                return col
    return None


def _construir_cache_merge(ws) -> dict[tuple[int, int], object]:
    cache: dict[tuple[int, int], object] = {}
    for rango in ws.merged_cells.ranges:
        valor_top_left = ws.cell(row=rango.min_row, column=rango.min_col).value
        for fila in range(rango.min_row, rango.max_row + 1):
            for col in range(rango.min_col, rango.max_col + 1):
                cache[(fila, col)] = valor_top_left
    return cache


def _valor_celda(ws, cache_merge: dict, fila: int, col: int) -> object:
    valor = ws.cell(row=fila, column=col).value
    if valor is not None:
        return valor
    return cache_merge.get((fila, col))


def _desambiguar_columnas(columnas: list[str]) -> list[str]:
    vistas: dict[str, int] = {}
    resultado = []
    for nombre in columnas:
        vistas[nombre] = vistas.get(nombre, 0) + 1
        if vistas[nombre] == 1:
            resultado.append(nombre)
        else:
            resultado.append(f"{nombre} ({vistas[nombre]})")
    return resultado


def listar_hojas(path) -> list[str]:
    wb = openpyxl.load_workbook(path, read_only=True)
    hojas = list(wb.sheetnames)
    wb.close()
    return hojas


def leer_hoja_aplanada(path, sheet_name: str, header_rows: int = 3) -> pd.DataFrame:
    """Lee una hoja completa y devuelve un DataFrame con columnas aplanadas.

    Todas las celdas quedan como texto/objeto crudo (sin castear tipos):
    la responsabilidad de parsear fechas/números es de src/normalizers.py,
    hoja por hoja, porque cada columna puede necesitar una heurística
    distinta (fecha vs texto que parece fecha, DNI numérico vs con ceros, etc).
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb[sheet_name]
        cache_merge = _construir_cache_merge(ws)
        max_col = ws.max_column
        max_row = ws.max_row

        columnas = []
        for col in range(1, max_col + 1):
            partes: list[str] = []
            for fila in range(1, header_rows + 1):
                valor = _valor_celda(ws, cache_merge, fila, col)
                texto = str(valor).strip() if valor is not None else ""
                texto = re.sub(r"\s+", " ", texto)
                if texto and texto not in partes:
                    partes.append(texto)
            nombre = " ".join(partes).strip()
            columnas.append(nombre if nombre else f"col_{col}")
        columnas = _desambiguar_columnas(columnas)

        filas = []
        indices_fila = []
        for fila in range(header_rows + 1, max_row + 1):
            valores = [ws.cell(row=fila, column=col).value for col in range(1, max_col + 1)]
            if all(v is None or (isinstance(v, str) and not v.strip()) for v in valores):
                continue
            filas.append(valores)
            indices_fila.append(fila)

        df = pd.DataFrame(filas, columns=columnas)
        df.insert(0, "_excel_row", indices_fila)
        return df
    finally:
        wb.close()
